"""
PDF Converter Service for extracting tables from PDFs and converting to Excel
"""
import os
import signal
import threading
from pathlib import Path
from typing import List, Optional, Tuple
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


class ConversionTimeoutError(Exception):
    """Exception raised when conversion takes too long"""
    pass


class TimeoutContext:
    """Context manager for timeout operations"""
    
    def __init__(self, seconds: int):
        self.seconds = seconds
        self.timer = None
        self.timed_out = False
    
    def _timeout_handler(self):
        """Handler called when timeout occurs"""
        self.timed_out = True
    
    def __enter__(self):
        """Start the timeout timer"""
        if self.seconds > 0:
            self.timer = threading.Timer(self.seconds, self._timeout_handler)
            self.timer.start()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Cancel the timeout timer"""
        if self.timer:
            self.timer.cancel()
        
        # If we timed out, raise the exception
        if self.timed_out:
            raise ConversionTimeoutError(f"Operation timed out after {self.seconds} seconds")
        
        return False


class PDFToExcelConverterService:
    """Service for converting PDF files to Excel format"""
    
    # Default timeout for conversion operations (in seconds)
    DEFAULT_TIMEOUT = 60
    
    def __init__(self, timeout: int = DEFAULT_TIMEOUT):
        """
        Initialize PDFToExcelConverterService
        
        Args:
            timeout: Maximum time in seconds for conversion operations
        """
        self.timeout = timeout
    
    def extract_tables_from_pdf(self, pdf_path: str) -> List[pd.DataFrame]:
        """
        Extract tables from PDF using pdfplumber
        
        Handles PDFs with multiple pages and multiple tables per page.
        Includes timeout mechanism to prevent hanging on large/complex PDFs.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            List of pandas DataFrames, one for each table found
            
        Raises:
            FileNotFoundError: If PDF file doesn't exist
            ValueError: If PDF cannot be opened or parsed
            ConversionTimeoutError: If extraction takes longer than timeout
            
        Requirements: 2.1, 2.2, 2.4
        """
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        all_tables = []
        
        try:
            # Use timeout context to prevent hanging
            with TimeoutContext(self.timeout) as timeout_ctx:
                with pdfplumber.open(pdf_path) as pdf:
                    # Check if PDF is valid and has pages
                    if not pdf.pages:
                        raise ValueError("PDF file has no pages")
                    
                    # Iterate through all pages
                    for page_num, page in enumerate(pdf.pages, start=1):
                        # Check for timeout periodically
                        if timeout_ctx.timed_out:
                            raise ConversionTimeoutError(f"Extraction timed out after {self.timeout} seconds")
                        
                        try:
                            # Extract tables from the page
                            tables = page.extract_tables()
                            
                            if tables:
                                for table_num, table in enumerate(tables, start=1):
                                    # Convert table to DataFrame
                                    if table and len(table) > 0:
                                        # First row as header if it exists
                                        if len(table) > 1:
                                            df = pd.DataFrame(table[1:], columns=table[0])
                                        else:
                                            # Single row table
                                            df = pd.DataFrame(table)
                                        
                                        # Clean up the DataFrame
                                        df = self._clean_dataframe(df)
                                        
                                        if not df.empty:
                                            all_tables.append(df)
                        except Exception as e:
                            # Log error but continue with other pages
                            print(f"Warning: Error extracting tables from page {page_num}: {str(e)}")
                            continue
        
        except pdfplumber.pdfminer.pdfparser.PDFSyntaxError:
            raise ValueError("PDF file is corrupted or invalid")
        except ConversionTimeoutError:
            # Re-raise timeout errors
            raise
        except Exception as e:
            # Check if it's a PDF parsing error
            if "PDF" in str(e) or "parse" in str(e).lower():
                raise ValueError(f"Failed to parse PDF file: {str(e)}")
            raise ValueError(f"Failed to process PDF file: {str(e)}")
        
        return all_tables
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean up extracted DataFrame
        
        Args:
            df: DataFrame to clean
            
        Returns:
            Cleaned DataFrame
        """
        # Create a copy to avoid modifying the original
        df = df.copy()
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Strip whitespace from string columns before filling NA
        for col in df.columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Replace None with empty string
        df = df.fillna('')
        
        return df

    def convert_dataframes_to_excel(self, dataframes: List[pd.DataFrame], output_path: str) -> bool:
        """
        Convert list of DataFrames to Excel file with multiple sheets
        
        Args:
            dataframes: List of pandas DataFrames to convert
            output_path: Path where Excel file should be saved
            
        Returns:
            True if conversion successful, False otherwise
            
        Raises:
            ValueError: If dataframes list is empty or invalid
            
        Requirements: 2.2, 2.3, 2.4
        """
        if not dataframes:
            raise ValueError("No dataframes provided for conversion")
        
        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for idx, df in enumerate(dataframes, start=1):
                    if df.empty:
                        continue
                    
                    # Create sheet name
                    sheet_name = f"Table_{idx}"
                    
                    # Write DataFrame to Excel
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the worksheet to apply formatting
                    worksheet = writer.sheets[sheet_name]
                    
                    # Apply formatting to preserve table structure
                    self._format_worksheet(worksheet, df)
            
            return True
        
        except PermissionError:
            raise ValueError(f"Permission denied: Cannot write to {output_path}")
        except Exception as e:
            raise ValueError(f"Error converting to Excel: {str(e)}")
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Apply formatting to Excel worksheet to preserve table structure
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame that was written to the worksheet
            
        Requirements: 2.3
        """
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Center align all cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def convert_pdf_to_excel(self, pdf_path: str, output_path: str) -> Tuple[bool, Optional[str]]:
        """
        Convert PDF file to Excel format
        
        This is the main conversion method that combines extraction and Excel generation.
        Handles PDFs without tables gracefully and provides detailed error messages.
        
        Args:
            pdf_path: Path to the PDF file
            output_path: Path where Excel file should be saved
            
        Returns:
            Tuple of (success, error_message)
            
        Requirements: 2.1, 2.2, 2.3, 2.4
        """
        try:
            # Validate input file exists
            if not os.path.exists(pdf_path):
                return False, f"PDF file not found: {pdf_path}"
            
            # Validate input is a file
            if not os.path.isfile(pdf_path):
                return False, f"Path is not a file: {pdf_path}"
            
            # Extract tables from PDF
            try:
                dataframes = self.extract_tables_from_pdf(pdf_path)
            except FileNotFoundError as e:
                return False, str(e)
            except ValueError as e:
                # Handle corrupted or invalid PDFs
                return False, f"Invalid PDF file: {str(e)}"
            except ConversionTimeoutError:
                return False, "Conversion timed out - PDF is too large or complex"
            
            # Handle PDFs without tables gracefully
            if not dataframes:
                return False, "No tables found in PDF"
            
            # Convert to Excel
            try:
                self.convert_dataframes_to_excel(dataframes, output_path)
            except ValueError as e:
                return False, str(e)
            
            # Verify output file was created
            if not os.path.exists(output_path):
                return False, "Failed to create Excel file"
            
            return True, None
        
        except Exception as e:
            # Catch any unexpected errors
            return False, f"Unexpected error during conversion: {str(e)}"
